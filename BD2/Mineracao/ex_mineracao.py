from datetime import datetime
import psycopg2
import openpyxl

combinacoes = {1:{}, 2:{}, 3:{}, 4:{}, 5:{}, 6:{}, 7:{}}
total_vendido = {}
itens = {}
lista_saida = [(),]
conn = None
nomes_colunas = None

#
#	Faz a consulta na base de dados e grava em uma lista de tuplas
#

try:
	conn = psycopg2.connect(dbname="postgres", user="postgres", password="postgres")
	cur = conn.cursor()
	cur.execute("SELECT * FROM vendas")
	nomes_colunas = [desc[0] for desc in cur.description]
	result = cur.fetchall()

except (Exception, psycopg2.DatabaseError) as error:
    print(error)
finally:
    if conn is not None:
        conn.close()

#
#	Lê as combinações
#

for row in range(len(result)):
	for index_base in range(8):
		for index_percorrido in range(8):
			if ((index_base>0) and (index_percorrido>0) and (index_base != index_percorrido)):
				if(result[row][index_base] and result[row][index_percorrido]):
					if index_percorrido not in combinacoes[index_base]:
						combinacoes[index_base][index_percorrido] = 0
					combinacoes[index_base][index_percorrido] += 1
					pass
				else:
					continue

#
#	Conta o numero total de registros
#

total_de_registros = len(result)

#
#	Conta o numero total de registro com item comprado
#

for index in range(8):
	for row in range(len(result)):
		if (result[row][index] and index > 0):
			if index not in total_vendido:
				total_vendido[index] = 0
			total_vendido[index] += 1

#
#	Pega os nomes dos itens
#

for x in range(1,8,1):
	itens[x] = nomes_colunas[x]

#
#	Calcula e guarda o suporte e o grau de certeza em uma lista de tuplas
#

contador_linhas = 0

for x in range(1,8,1):
	for y in range(1,8,1):
		if((x != y) and y in combinacoes[x]):
			sup = combinacoes[x][y]/total_de_registros
			conf = combinacoes[x][y]/total_vendido[x]
			lista_saida.append((itens[x] , itens[y] , combinacoes[x][y] , sup , conf))
			contador_linhas+=1

#
#	Grava a lista de tuplas em uma planilha .xlsx
#

try:
	nome_arquivo = 'resultado.xlsx'
	arquivo = openpyxl.Workbook()
	planilha = arquivo.active
	planilha.title = 'resultado'
	planilha.append(("item X", "item Y", "vendas" ,"sup", "conf"))
	for linha in lista_saida:
	    planilha.append(linha)
	planilha.cell(row=contador_linhas +4, column = 1).value = datetime.now().strftime('%d/%m/%Y')
	planilha.cell(row=contador_linhas +4, column = 2).value = datetime.now().strftime('%H:%M:%S')
	arquivo.save(nome_arquivo)
	print("\n** Os resultados foram gravados no arquivo",nome_arquivo," **")
	arquivo.close()
except (Exception, openpyxl.utils.exceptions.ReadOnlyWorkbookException) as error:
	print(error)